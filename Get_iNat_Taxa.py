#!/usr/bin/ python3
# -*- coding: utf-8 -*-
"""
Created on Thu Dec 30 16:03:53 2021

@author: manuel
"""
filepath_parent = "/home/manuel/Downloads"
filename = "Veg_Lists_manual.xlsx"

import os
filepath = os.path.join(filepath_parent, filename)
project_ID = "lele-herbs-and-grasses"

from pyinaturalist import *
observations = []
page_results = [1]
p = 1
while len(page_results) > 0:
    page = get_observations(project_id = project_ID, per_page = 30, page = p)
    page_results = page_results["results"]
    observations += page_results
    p += 1

iNat_data = list()
iNat_notes = list()
for obs in observations:
    if obs["description"] is not None and obs["description"] != "":
        note = obs["description"]
        taxon = obs["taxon"]["name"]
        url = obs["uri"]
        quality = obs["quality_grade"]
        iNat_data.append([taxon, url, quality])
        iNat_notes.append(note)

# edit workbook
from openpyxl import Workbook, load_workbook

wb = load_workbook(filepath)
sht0 = wb["manual_veg_list"]
c_names = [cell.value for cell in sht0[1]]

# add missing columns
def add_col(colnames, colname, varname):
    for c, i in zip(colnames, range(len(colnames))):
        if c == colname:
            globals()[varname] = i + 1
    if varname not in globals():
        new_col_idx = sht0.max_column + 1
        sht0.cell(row = 1, column = new_col_idx).value = colname
        globals()[varname] = new_col_idx

for c_name, v_name in zip(["iNaturalist_ID", "iNaturalist_taxon", \
                           "iNaturalist_url", "iNaturalist_status"], \
                          ["c_iNatID", "c_iNatTax", \
                           "c_iNatURL", "c_iNatStat"]):
    add_col(colnames = c_names, colname = c_name, varname = v_name)

# check by row
for r in range(2, sht0.max_row + 1):
    iNaturalist_ID = sht0.cell(row = r, column = c_iNatID).value
    if iNaturalist_ID is not None:
        iNaturalist_ID = str(iNaturalist_ID)
        matches = []
        for note in iNat_notes:
            if iNaturalist_ID in str(note):
                matches.append(note)
        if len(matches) < 1:
            sht0.cell(row = r, column = c_iNatStat).value = "No matching " + \
                "iNaturalist observation found."
        elif len(matches) == 1:
            data = iNat_data[iNat_notes.index(matches[0])]
            print("r = " + str(r) + "note: " + note, " index: " + str(iNat_notes.index(note)) +\
                  "iNatID: " + iNaturalist_ID)
            sht0.cell(row = r, column = c_iNatTax).value = data[0]
            sht0.cell(row = r, column = c_iNatURL).value = data[1]
            sht0.cell(row = r, column = c_iNatStat).value = data[2]
        else:
            data = iNat_data[iNat_notes.index(note)]
            sht0.cell(row = r, column = c_iNatTax).value = data[0]
            sht0.cell(row = r, column = c_iNatURL).value = data[1]
            sht0.cell(row = r, column = c_iNatStat).value = "iNaturalist" + \
                "_ID ambiguous"
wb.save(filepath)
wb.close()
