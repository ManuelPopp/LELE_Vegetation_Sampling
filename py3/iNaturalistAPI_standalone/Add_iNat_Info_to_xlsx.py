#!/usr/bin/ python3
# -*- coding: utf-8 -*-
"""
Created on Thu Dec 30 16:03:53 2021

@author: manuel
"""
# path to folder containing the Excel sheet
filepath_parent = "/home/manuel/Downloads"
# Excel sheet file name
filename = "Vegetation List ASICS project.xlsx"
# Name of sheet containing the species-level data
sheet_name = "List of species"
# Name of column contining the IDs to connect to iNaturalist observations
ID_field = "Notes"
# Project ID (last part of URL to iNaturalist project)
project_ID = "flora-fauna-of-the-sani-pass-road"
# Format of the ID used in the iNaturalist observation comments as RegEx
# hint: \d = any integer; \w = any letter/symbol (excluding line break);
# . = any symbol
ID_format = "S\d\w.\d\d\d"

ID_format = ID_format.casefold()

import os
filepath = os.path.join(filepath_parent, filename)

from pyinaturalist import *
observations = []
page_results = [1]
p = 1
while len(page_results) > 0:
    page = get_observations(project_id = project_ID, per_page = 30, page = p)
    page_results = page["results"]
    observations += page_results
    p += 1

iNat_data = list()
iNat_notes = list()
for obs in observations:
    if obs["description"] is not None and obs["description"] != "":
        note = obs["description"]
        if obs["taxon"] is not None:
            taxon = obs["taxon"]["name"]
        else:
            taxon = None
        url = obs["uri"]
        quality = obs["quality_grade"]
        iNat_data.append([taxon, url, quality])
        iNat_notes.append(note)

# edit workbook
from openpyxl import Workbook, load_workbook

wb = load_workbook(filepath)
sht0 = wb[sheet_name]
c_names = [cell.value for cell in sht0[1]]

# add missing columns
def add_col(colnames, colname, varname):
    for c, i in zip(colnames, range(len(colnames))):
        if c == colname:
            globals()[varname] = i + 1
    if varname not in globals():
        new_col_idx = sht0.max_column + 1
        sht0.cell(row = 1, column = new_col_idx).value = colname
        globals()[varname] = new_col_idx

for c_name, v_name in zip(["iNaturalist_ID", "iNaturalist_taxon", \
                           "iNaturalist_url", "iNaturalist_status"], \
                          ["c_iNatID", "c_iNatTax", \
                           "c_iNatURL", "c_iNatStat"]):
    add_col(colnames = c_names, colname = c_name, varname = v_name)

# check by row
import re
regEx = re.compile(ID_format)

c = c_names.index(ID_field) + 1
for r in range(2, sht0.max_row + 1):
    print("Searching for matching IDs", str(r - 1), "of", str(sht0.max_row))
    iNaturalist_ID_value = sht0.cell(row = r, column = c).value
    if iNaturalist_ID_value is not None:
        if regEx.search(str(iNaturalist_ID_value).casefold()) is not None:
            iNaturalist_ID = regEx.search(str(iNaturalist_ID_value)\
                                          .casefold())[0]
            matches = []
            for note in iNat_notes:
                if iNaturalist_ID in str(note).casefold():
                    matches.append(note)
            if len(matches) < 1:
                sht0.cell(row = r,\
                          column = c_iNatStat).value = "No matching " + \
                    "iNaturalist observation found."
            elif len(matches) == 1:
                data = iNat_data[iNat_notes.index(matches[0])]
                print("r = " + str(r) + "note: " + note, " index: " + \
                      str(iNat_notes.index(note)) + \
                      "iNatID: " + iNaturalist_ID)
                sht0.cell(row = r, column = c_iNatID).value = iNaturalist_ID
                sht0.cell(row = r, column = c_iNatTax).value = data[0]
                sht0.cell(row = r, column = c_iNatURL).value = data[1]
                sht0.cell(row = r, column = c_iNatStat).value = data[2]
            else:
                data = iNat_data[iNat_notes.index(note)]
                sht0.cell(row = r, column = c_iNatID).value = iNaturalist_ID
                sht0.cell(row = r, column = c_iNatTax).value = data[0]
                sht0.cell(row = r, column = c_iNatURL).value = data[1]
                sht0.cell(row = r, column = c_iNatStat)\
                    .value = "iNaturalist" + "_ID ambiguous"
wb.save(filepath)
wb.close()
